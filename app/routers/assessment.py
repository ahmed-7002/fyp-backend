import uuid
from datetime import datetime, timedelta, timezone
from io import BytesIO
from typing import Literal

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.auth import get_current_user
from app.data.dass21_questions import DASS21_QUESTIONS, ANSWER_LABELS
from app.database import get_db
from app.models import Assessment
from app.schemas import (
    AssessmentInsight,
    AssessmentSubmitIn,
    AssessmentSubmitOut,
    AssessmentSummary,
    DassResult,
    FerResult,
)
from app.services.dass_scoring import score_dass21
from app.services.risk_engine import compute_overall_risk, generate_actionable_tips

router = APIRouter(prefix="/api", tags=["assessment"])

# Maps the profile page's time-range filter keys to a day count. "all" is
# handled separately below (no cutoff applied at all).
_INSIGHTS_RANGE_DAYS = {"7d": 7, "30d": 30, "90d": 90}

# Same teal used for the header row/chart accents in the old frontend
# exceljs export (and the app's --color-teal-dark), so the downloaded
# workbook still matches the app's look now that generation has moved
# server-side.
_HEADER_FILL = PatternFill(start_color="FF2B6F6B", end_color="FF2B6F6B", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_NOTE_FONT = Font(bold=True, color="FF1E4F4C")


def _style_header_row(ws: Worksheet, row: int = 1) -> None:
    """Bolds + teal-fills a worksheet's header row, mirroring the
    headerRow.font/headerRow.fill styling the frontend used to apply via
    exceljs."""
    for cell in ws[row]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _build_response(
    record: Assessment,
    dass_result: DassResult | None,
    fer_result: FerResult | None,
    tips: list[dict],
) -> AssessmentSubmitOut:
    """Shared response shape for both POST /api/assessments (fresh
    submission) and GET /api/assessments/{id} (re-viewed later) - was
    previously duplicated identically in both endpoints."""
    return AssessmentSubmitOut(
        id=record.id,
        dass_result=dass_result,
        fer_result=fer_result,
        final_risk_level=record.final_risk_level,
        final_summary=record.final_summary,
        actionable_tips=tips,
        created_at=record.created_at,
    )


@router.get("/dass21/questions")
def get_questions():
    return {
        "answer_labels": ANSWER_LABELS,
        "questions": [{"number": i + 1, "text": q} for i, q in enumerate(DASS21_QUESTIONS)],
    }


@router.post("/assessments", response_model=AssessmentSubmitOut)
def submit_assessment(
    payload: AssessmentSubmitIn,
    db: Session = Depends(get_db),
    user_id: str = Depends(get_current_user),
):
    if payload.clerk_user_id != user_id:
        raise HTTPException(403, "clerk_user_id does not match authenticated user")

    needs_dass = payload.assessment_mode in ("questionnaire", "combined")
    needs_fer = payload.assessment_mode == "video"

    if needs_dass and not payload.dass_answers:
        raise HTTPException(400, "dass_answers is required for this assessment mode")
    if needs_fer and not payload.fer_result:
        raise HTTPException(400, "fer_result is required for this assessment mode")

    dass_result: DassResult | None = None
    if payload.dass_answers:
        dass_result = score_dass21(payload.dass_answers)

    fer_result: FerResult | None = payload.fer_result

    # Age and gender both come straight from the onboarding step of the
    # submission payload (the same values that already get persisted onto
    # the Assessment record below) - there's no separate lookup needed
    # since the user provides both as part of every submission. Age is cast
    # defensively in case it ever arrives as a numeric string from an older
    # client; gender is passed through as-is (risk_engine normalises it,
    # and treats "Prefer not to say" / anything unrecognised as unknown).
    age: int | None = int(payload.age) if payload.age is not None else None
    gender: str | None = payload.gender

    # compute_overall_risk now returns a 3-tuple - risk level, summary, and
    # the rule-based bilingual tips generated from these same DASS/FER
    # values plus age and gender. Tips are NOT stored on `record` below and
    # NOT part of models.py/schema.sql - they're only ever attached to the
    # API response. Neither `age` nor `gender` ever affects
    # final_risk_level/final_summary - both only affect tip WORDING (see
    # risk_engine.py docstring).
    risk_level, summary, tips = compute_overall_risk(dass_result, fer_result, age, gender)

    record = Assessment(
        id=uuid.uuid4(),
        clerk_user_id=payload.clerk_user_id,
        full_name=payload.full_name,
        # Use the same normalised `age`/`gender` locals that were just fed
        # into compute_overall_risk() above, not payload.age/payload.gender
        # directly - keeps what's persisted and what generated these tips
        # identical, including the defensive int-cast on age.
        age=age,
        gender=gender,
        assessment_mode=payload.assessment_mode,
        final_risk_level=risk_level,
        final_summary=summary,
    )

    if payload.dass_answers:
        for i, ans in enumerate(payload.dass_answers, start=1):
            setattr(record, f"dass_q{i}", ans)
        record.dass_depression_score = dass_result.depression_score
        record.dass_anxiety_score = dass_result.anxiety_score
        record.dass_stress_score = dass_result.stress_score
        record.dass_depression_severity = dass_result.depression_severity
        record.dass_anxiety_severity = dass_result.anxiety_severity
        record.dass_stress_severity = dass_result.stress_severity

    if fer_result:
        record.fer_frames_captured = fer_result.frames_captured
        record.fer_frames_analyzed = fer_result.frames_analyzed
        record.fer_angry = fer_result.angry
        record.fer_disgust = fer_result.disgust
        record.fer_fear = fer_result.fear
        record.fer_happy = fer_result.happy
        record.fer_sad = fer_result.sad
        record.fer_surprise = fer_result.surprise
        record.fer_neutral = fer_result.neutral
        record.fer_dominant_emotion = fer_result.dominant_emotion

    db.add(record)
    db.commit()
    db.refresh(record)

    return _build_response(record, dass_result, fer_result, tips)


@router.get("/assessments", response_model=list[AssessmentSummary])
def list_assessments(
    db: Session = Depends(get_db),
    user_id: str = Depends(get_current_user),
):
    """Returns every past session for the signed-in user, most recent first.

    Note: anonymized (deleted) sessions have clerk_user_id set to NULL,
    so they never match this filter and correctly stop appearing here
    for anyone, without needing any extra "is_deleted" flag."""
    records = (
        db.query(Assessment)
        .filter(Assessment.clerk_user_id == user_id)
        .order_by(Assessment.created_at.desc())
        .all()
    )
    return records


def _fetch_insight_rows(
    db: Session,
    user_id: str,
    range_key: Literal["7d", "30d", "90d", "all"],
) -> list[AssessmentInsight]:
    """Range-filtered, column-trimmed data for the profile page's Insights
    charts (Overall Progress + Facial Emotion Summary), and now also for
    the server-side Excel export (GET /assessments/export/excel) - both
    need the exact same rows, so this is the one place that query lives.

    Replaces the previous frontend pattern of calling
    GET /api/assessments/{id} once per session - an N+1 fan-out that also
    pulled every column (all 21 raw dass_qN answers, full_name,
    final_summary, actionable_tips) for every past session on every
    profile load, none of which the charts (or the export) use.

    Two things keep this fast as history grows:
    - The date-range cutoff is applied in the SQL WHERE clause, not by
      fetching every session and filtering client-side.
    - Only the ~15 columns the charts actually read are selected (see
      AssessmentInsight), so the heavy columns above never leave the DB.

    A composite index on (clerk_user_id, created_at) is recommended so
    this stays a fast index scan rather than a sequential scan as history
    grows.
    """
    query = db.query(
        Assessment.id,
        Assessment.created_at,
        Assessment.assessment_mode,
        Assessment.final_risk_level,
        Assessment.dass_depression_score,
        Assessment.dass_anxiety_score,
        Assessment.dass_stress_score,
        Assessment.dass_depression_severity,
        Assessment.dass_anxiety_severity,
        Assessment.dass_stress_severity,
        Assessment.fer_frames_captured,
        Assessment.fer_frames_analyzed,
        Assessment.fer_angry,
        Assessment.fer_disgust,
        Assessment.fer_fear,
        Assessment.fer_happy,
        Assessment.fer_sad,
        Assessment.fer_surprise,
        Assessment.fer_neutral,
        Assessment.fer_dominant_emotion,
    ).filter(Assessment.clerk_user_id == user_id)

    if range_key != "all":
        cutoff = datetime.now(timezone.utc) - timedelta(days=_INSIGHTS_RANGE_DAYS[range_key])
        query = query.filter(Assessment.created_at >= cutoff)

    rows = query.order_by(Assessment.created_at.asc()).all()

    results: list[AssessmentInsight] = []
    for row in rows:
        dass_result = None
        if row.dass_depression_score is not None:
            dass_result = DassResult(
                depression_score=row.dass_depression_score,
                anxiety_score=row.dass_anxiety_score,
                stress_score=row.dass_stress_score,
                depression_severity=row.dass_depression_severity,
                anxiety_severity=row.dass_anxiety_severity,
                stress_severity=row.dass_stress_severity,
            )

        fer_result = None
        if row.fer_frames_analyzed is not None:
            fer_result = FerResult(
                frames_captured=row.fer_frames_captured,
                frames_analyzed=row.fer_frames_analyzed,
                angry=row.fer_angry,
                disgust=row.fer_disgust,
                fear=row.fer_fear,
                happy=row.fer_happy,
                sad=row.fer_sad,
                surprise=row.fer_surprise,
                neutral=row.fer_neutral,
                dominant_emotion=row.fer_dominant_emotion,
            )

        results.append(
            AssessmentInsight(
                id=row.id,
                created_at=row.created_at,
                assessment_mode=row.assessment_mode,
                final_risk_level=row.final_risk_level,
                dass_result=dass_result,
                fer_result=fer_result,
            )
        )

    return results


def _progress_rows(rows: list[AssessmentInsight]) -> list[tuple[str, str, int, int, int]]:
    """Mirrors Profile.jsx's old `progressData` derivation: only
    questionnaire/combined sessions that carry a DASS-21 result, one row
    per session with Date, Time, Depression, Anxiety, Stress. `rows` is
    already oldest-first (see _fetch_insight_rows' ORDER BY), so the
    output - and therefore the exported "Trend Data" sheet and its line
    chart - reads chronologically top to bottom."""
    out: list[tuple[str, str, int, int, int]] = []
    for r in rows:
        if r.assessment_mode not in ("questionnaire", "combined") or r.dass_result is None:
            continue
        created = r.created_at
        out.append(
            (
                created.strftime("%b %d, %Y"),
                created.strftime("%I:%M %p"),
                r.dass_result.depression_score,
                r.dass_result.anxiety_score,
                r.dass_result.stress_score,
            )
        )
    return out


def _emotion_rows(rows: list[AssessmentInsight]) -> list[tuple[str, int]]:
    """Mirrors Profile.jsx's old `emotionData` derivation: counts how many
    video/combined sessions had each emotion as their dominant one."""
    counts: dict[str, int] = {}
    for r in rows:
        if r.assessment_mode not in ("video", "combined"):
            continue
        if r.fer_result is None or r.fer_result.frames_analyzed <= 0:
            continue
        emo = r.fer_result.dominant_emotion
        counts[emo] = counts.get(emo, 0) + 1
    return [(emo.capitalize(), count) for emo, count in counts.items()]


# NOTE: this must stay registered ABOVE GET /assessments/{assessment_id}.
# FastAPI matches routes in declaration order, and "insights" would
# otherwise be swallowed by the {assessment_id} path (and fail UUID
# validation) before ever reaching this one.
@router.get("/assessments/insights", response_model=list[AssessmentInsight])
def get_assessments_insights(
    range: Literal["7d", "30d", "90d", "all"] = "30d",
    db: Session = Depends(get_db),
    user_id: str = Depends(get_current_user),
):
    """Range-filtered, column-trimmed data for the profile page's Insights
    charts (Overall Progress + Facial Emotion Summary). See
    _fetch_insight_rows for the query itself, which this now just calls
    directly - it's also reused by GET /assessments/export/excel below."""
    return _fetch_insight_rows(db, user_id, range)


# Also registered above GET /assessments/{assessment_id} for the same
# routing reason as /assessments/insights above: "/assessments/export/excel"
# is a distinct two-segment path so it wouldn't actually collide with the
# single-segment {assessment_id} route either way, but keeping every
# static Assessment sub-route declared before the catch-all UUID route
# keeps this file's ordering easy to reason about.
@router.get("/assessments/export/excel")
def export_assessments_excel(
    range: Literal["7d", "30d", "90d", "all"] = "30d",
    db: Session = Depends(get_db),
    user_id: str = Depends(get_current_user),
):
    """Builds a .xlsx with the same range-filtered data as
    GET /assessments/insights, using openpyxl instead of the old frontend
    exceljs flow.

    Three sheets:
    - "Trend Data": one row per questionnaire/combined session (Date,
      Time, Depression, Anxiety, Stress) - the raw numbers behind the
      Overall Progress chart.
    - "Emotion Data": one row per dominant emotion (Emotion, Count) - the
      raw numbers behind the Facial Emotion Summary chart.
    - "Charts": a real openpyxl LineChart and BarChart, built with
      Reference()s that point at the two sheets above rather than any
      pasted-in image. That's the whole point of moving this
      server-side - opening the download in Excel and editing a score in
      "Trend Data" redraws the line chart on "Charts", which a rasterized
      PNG (the old frontend approach) could never do.

    Reuses _fetch_insight_rows so the numbers in this download always
    match what the profile page's own Insights charts are currently
    showing for the same `range`.
    """
    insight_rows = _fetch_insight_rows(db, user_id, range)
    progress_rows = _progress_rows(insight_rows)
    emotion_rows = _emotion_rows(insight_rows)

    wb = Workbook()

    # --- Sheet 1: Trend Data ------------------------------------------
    trend_ws = wb.active
    trend_ws.title = "Trend Data"
    trend_ws.append(["Date", "Time", "Depression", "Anxiety", "Stress"])
    for row in progress_rows:
        trend_ws.append(list(row))
    _style_header_row(trend_ws)
    for col, width in zip("ABCDE", (14, 12, 13, 13, 13)):
        trend_ws.column_dimensions[col].width = width

    # --- Sheet 2: Emotion Data -----------------------------------------
    emotion_ws = wb.create_sheet("Emotion Data")
    emotion_ws.append(["Emotion", "Count"])
    for row in emotion_rows:
        emotion_ws.append(list(row))
    _style_header_row(emotion_ws)
    for col, width in zip("AB", (14, 10)):
        emotion_ws.column_dimensions[col].width = width

    # --- Sheet 3: Charts -------------------------------------------------
    # Built from Reference()s into the two sheets above, not images - these
    # stay live, editable Excel chart objects once opened.
    charts_ws = wb.create_sheet("Charts")

    if progress_rows:
        last_row = 1 + len(progress_rows)
        line_chart = LineChart()
        line_chart.title = "Overall Progress"
        line_chart.style = 2
        line_chart.y_axis.title = "Score"
        line_chart.x_axis.title = "Date"
        line_chart.width = 24
        line_chart.height = 11
        data = Reference(trend_ws, min_col=3, max_col=5, min_row=1, max_row=last_row)
        categories = Reference(trend_ws, min_col=1, min_row=2, max_row=last_row)
        line_chart.add_data(data, titles_from_data=True)
        line_chart.set_categories(categories)
        for series in line_chart.series:
            series.smooth = False
        charts_ws.add_chart(line_chart, "A1")
    else:
        charts_ws["A1"] = "No question-based sessions in this range yet."
        charts_ws["A1"].font = _NOTE_FONT

    if emotion_rows:
        last_row = 1 + len(emotion_rows)
        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.title = "Facial Emotion Summary"
        bar_chart.style = 10
        bar_chart.y_axis.title = "Sessions"
        bar_chart.x_axis.title = "Emotion"
        bar_chart.width = 24
        bar_chart.height = 11
        data = Reference(emotion_ws, min_col=2, min_row=1, max_row=last_row)
        categories = Reference(emotion_ws, min_col=1, min_row=2, max_row=last_row)
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(categories)
        charts_ws.add_chart(bar_chart, "A22")
    else:
        charts_ws["A22"] = "No video sessions in this range yet."
        charts_ws["A22"].font = _NOTE_FONT

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"mindful-checkin-insights-{range}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/assessments/{assessment_id}", response_model=AssessmentSubmitOut)
def get_assessment(
    assessment_id: uuid.UUID,
    db: Session = Depends(get_db),
    user_id: str = Depends(get_current_user),
):
    record = db.query(Assessment).filter(Assessment.id == assessment_id).first()
    if not record:
        raise HTTPException(404, "Assessment not found")
    if record.clerk_user_id != user_id:
        raise HTTPException(403, "Not authorized to view this assessment")

    dass_result = None
    if record.dass_depression_score is not None:
        dass_result = DassResult(
            depression_score=record.dass_depression_score,
            anxiety_score=record.dass_anxiety_score,
            stress_score=record.dass_stress_score,
            depression_severity=record.dass_depression_severity,
            anxiety_severity=record.dass_anxiety_severity,
            stress_severity=record.dass_stress_severity,
        )

    fer_result = None
    if record.fer_frames_captured is not None:
        fer_result = FerResult(
            frames_captured=record.fer_frames_captured,
            frames_analyzed=record.fer_frames_analyzed,
            angry=record.fer_angry,
            disgust=record.fer_disgust,
            fear=record.fer_fear,
            happy=record.fer_happy,
            sad=record.fer_sad,
            surprise=record.fer_surprise,
            neutral=record.fer_neutral,
            dominant_emotion=record.fer_dominant_emotion,
        )

    # Age and gender are both read directly off the already-persisted
    # record (they were stored at submission time above, straight from the
    # onboarding payload) - no extra lookup against the auth/user object is
    # needed here either.
    age: int | None = record.age
    gender: str | None = record.gender

    # Tips aren't stored - they're recomputed here from the same saved
    # DASS/FER/age/gender values, using the standalone helper directly
    # rather than compute_overall_risk() (which would needlessly recompute
    # final_risk_level/final_summary, both of which are already stored on
    # `record` and read directly below).
    tips = generate_actionable_tips(dass_result, fer_result, age, gender)

    return _build_response(record, dass_result, fer_result, tips)


@router.delete("/assessments/{assessment_id}", status_code=204)
def delete_assessment(
    assessment_id: uuid.UUID,
    db: Session = Depends(get_db),
    user_id: str = Depends(get_current_user),
):
    """
    "Deletes" a session via anonymization, not a hard SQL delete: the two
    personally-identifying columns (clerk_user_id, full_name) are wiped to
    NULL, while age, gender, every DASS-21 answer/score, and every FER
    metric are left fully intact.

    Note: `age` and `gender` are deliberately kept (not wiped) here, same
    as before - this means tip generation for an anonymized session still
    tailors by age/gender if it's ever re-viewed, since neither is
    personally identifying in the way clerk_user_id/full_name are.
    """
    record = db.query(Assessment).filter(Assessment.id == assessment_id).first()
    if not record:
        raise HTTPException(404, "Assessment not found")
    if record.clerk_user_id != user_id:
        raise HTTPException(403, "Not authorized to delete this assessment")

    record.clerk_user_id = None
    record.full_name = None
    db.commit()
    return None